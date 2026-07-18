#!/usr/bin/env python3
"""SDS Biocompatibility Screening Tool — ACONIS | Claude AI | ISO 10993-1:2025"""

import streamlit as st
import anthropic
import fitz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import column_index_from_string
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
import json, io, datetime, re, os, traceback, base64, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="SDS Biocompatibility Screener", page_icon="🔬",
                   layout="wide", initial_sidebar_state="expanded")

# ─── File paths ───────────────────────────────────────────────────────────────
DIR         = os.path.dirname(__file__)
DB_FILE     = os.path.join(DIR, "databases.xlsx")
TMPL_FILE   = os.path.join(DIR, "template.xlsx")
PROMPT_FILE = os.path.join(DIR, "agent_prompt_generic.txt")

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAMgAAAA8CAYAAAAjW/WRAAAACXBIWXMAAA7EAAAOxAGVKw4bAAAEwGlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0IGJlZ2luPSfvu78nIGlkPSdXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQnPz4KPHg6eG1wbWV0YSB4bWxuczp4PSdhZG9iZTpuczptZXRhLyc+CjxyZGY6UkRGIHhtbG5zOnJkZj0naHR0cDovL3d3dy53My5vcmcvMTk5OS8wMi8yMi1yZGYtc3ludGF4LW5zIyc+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczpBdHRyaWI9J2h0dHA6Ly9ucy5hdHRyaWJ1dGlvbi5jb20vYWRzLzEuMC8nPgogIDxBdHRyaWI6QWRzPgogICA8cmRmOlNlcT4KICAgIDxyZGY6bGkgcmRmOnBhcnNlVHlwZT0nUmVzb3VyY2UnPgogICAgIDxBdHRyaWI6Q3JlYXRlZD4yMDI1LTA2LTI2PC9BdHRyaWI6Q3JlYXRlZD4KICAgICA8QXR0cmliOkV4dElkPjgzZGQ2N2U4LTc4MzAtNGFhYi1iNWE0LTliYjhlOTIxZjdkMDwvQXR0cmliOkV4dElkPgogICAgIDxBdHRyaWI6RmJJZD41MjUyNjU5MTQxNzk1ODA8L0F0dHJpYjpGYklkPgogICAgIDxBdHRyaWI6VG91Y2hUeXBlPjI8L0F0dHJpYjpUb3VjaFR5cGU+CiAgICA8L3JkZjpsaT4KICAgPC9yZGY6U2VxPgogIDwvQXR0cmliOkFkcz4KIDwvcmRmOkRlc2NyaXB0aW9uPgoKIDxyZGY6RGVzY3JpcHRpb24gcmRmOmFib3V0PScnCiAgeG1sbnM6ZGM9J2h0dHA6Ly9wdXJsLm9yZy9kYy9lbGVtZW50cy8xLjEvJz4KICA8ZGM6dGl0bGU+CiAgIDxyZGY6QWx0PgogICAgPHJkZjpsaSB4bWw6bGFuZz0neC1kZWZhdWx0Jz5EZXNpZ24gc2FucyB0aXRyZSAtIDE8L3JkZjpsaT4KICAgPC9yZGY6QWx0PgogIDwvZGM6dGl0bGU+CiA8L3JkZjpEZXNjcmlwdGlvbj4KCiA8cmRmOkRlc2NyaXB0aW9uIHJkZjphYm91dD0nJwogIHhtbG5zOnBkZj0naHR0cDovL25zLmFkb2JlLmNvbS9wZGYvMS4zLyc+CiAgPHBkZjpBdXRob3I+RWxvZGllIFNhdWRyYWlzPC9wZGY6QXV0aG9yPgogPC9yZGY6RGVzY3JpcHRpb24+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczp4bXA9J2h0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC8nPgogIDx4bXA6Q3JlYXRvclRvb2w+Q2FudmEgKFJlbmRlcmVyKSBkb2M9REFHcmVoMFBsVjggdXNlcj1VQUN3Zk1iN0hfTSBicmFuZD1CQUN3ZlBmcnRRZyB0ZW1wbGF0ZT08L3htcDpDcmVhdG9yVG9vbD4KIDwvcmRmOkRlc2NyaXB0aW9uPgo8L3JkZjpSREY+CjwveDp4bXBtZXRhPgo8P3hwYWNrZXQgZW5kPSdyJz8+8WcdTwAAExRJREFUeJztnXd0FdXWwH+TToKhJxDAAJFeDFXpTbooTxBFFBUR9PEen+Up6tPP51JULA+figVUBFEeooAFKSJV6SUghN5LgBAglBBS7vn+2OS7M3Pnzp17QxHX/Na6a2XqOTNz9jl777P3iaaUUri4uFgSdrUr4OLyR8YVEBcXG1wBcXGxwRUQFxcbXAFxcbHBFRAXFxtcAfkTU1AIaRmwLRNcZ35oRFztCrhcHpbth3umwr5Tst2iCnx7D1QpdXXrda3hjiB/QpSCoTO8wgGw6iB8uPLq1elaxRWQPyln83x1qlO5V6Ei1ziugPwJ0TQY1kJD0+2LCoe+9a9ala5ZNDcW68+JUvDrPvh2M0RHwKBUqJ94tWt17eEKiIuLDa6K5eJigysgLi42uALick3hUc4mPVUQ59pR7IlCpaBQgQaEhWHwnFwOrnR5IC8aIOxKFKaj0HOx3Cv0nFYodbGhAeGaeMiuBgezYdhMWLJXPHJ3N4I3e0BspO+5y/fD8O8h/RiUjYXnOsDwm0Kre9BG+r6T8NN2CV9YdQjWHISCix+yZJSiYw2NBonQoCL0qQslLB4gGHILYGY6bDgC6w7JC7pQoACNhJLQ9QZoWBE6pygaV9KK/QE9ChbthuUHYMsxWLAbjpyRY0nxio7VNeolQsvroV214IRmwlrYn+3dviUFWid7t8/nw1cbFKsPavywFTLOyHNWvA561oIbK0G/BlDpusBlHTkDH6/2bmvAC50CC9qFAvh+C2w8ChszFAt3axfnVDRiI6FzCtyYBE0rQa/aEBHu/PlDpcADN74nDV7PrXXgu3uN32BDBtz0IVwoNJ770e0wrEXwZTsSEKUkdGHUIpi7w9ujBiI+WiR9dDcoXSK4ip06D8/Ph/9uhKwcZ9e0uh5e6AjdagbfWxR4YPxqGPMb7Mhydk3dCjCileLhZhrhDpTVtuMUv+7zVuyN7vBUW8gvhJcXKj5apZF5zv4e4WEwpBmM6Wnf+WzIgNT3vdthGhS84l9AzufDc/Ngchocd/i+q5aCwU2lh466jIIydSPcPdX62Cd/gYeayd+5BdDkfdiS6Xtew0TYOCL4sgN+1rxC+NsPinbjYfZ258IBcPoCjFsN1d6En7Y5u0YpEYrkN2HsCufCASLEPSbCvdPgbJ7z6zYdhZs/VPz1e+fCAfIhHv1Oo804xc4grtOz/xS0/AheXhhYOEDUro9XSePfezK0Ms1sOgq1x8A7y5wLB8CBbHhpATQdC7tPXJq6WGEXATByLuTky9/jV1sLB8CJ86GVbSsgBR7oPQk+WKlZCoaGIjYSSsUo4qMhxo9Fk30Bbp0E32yyr4xSMHKuYsBUES4roiOgVAxcF43fXvurDdB+PJxw8LF/2QWtP4a1h6371nBNUTJK1MdwP93vigMajd5V/Lo3cHl69p6EWz6DtYeDuw5g+3HoOdHZM9qx5ZjU4UC27zENKB0DqRWhaRKUKWE9Am06Ch0/sb7HpeCWFP8jVFYOTNkgbWfcav+9d5tkv4dssTXSX18M83Ya95UtoXiijUbnGtCwokZcFOhfW+Y5CbGenKaYsgHyPXJMAQ9Nh5ur+o8oHbUI3lzq+wlurASPt5KHrFHWqz7leyDtMMzaJqqRXqjWHYY7p8CcByDSz8tddQDu+NJXGBNLKp5pr9GhOjRMLFKfNPI9sOmICNXoJcbe9nyBRq9J8POD0KKqdXlmPlplHJETS4qK2Pp6qF0BIsLg0GnYngUT18nIqj9/SyY8OVsxoW9ohpcChv8AR89694Vp0KcePNICbqoC8THGazLPwqI98q02HPHu358ND34DPw8uviGfVyiCe6EA6iVASjl4tSs8NVvqbObbTXBTVdh81LrgKqXgnV4iRAezYecJuW9CycA2mV8b5NhZUXNyC7z7utygmDZAo1SM1RW+LNqtuH2yZmiAj94EH9zme+7iPdDpU2MDKBUNY28TOyaQjn/qPNz/DXy/1bj/7R7wRBuL83Oh7hg4cta4/x9t4JUuMlLZcaEAnpkL/1lm/Gj1EuC3YdLzmjHbIEVowBOt4dVu9rr8L7ug31dGlSMyDDKeVZSLNd7XiQ2Sfgzq/8dYxtS74c4GgRt5gQcengGfrzM+x7bHFTXLhy4hu0/ArZMUWzLlHuVi4av+0OUG+GAlPD5LOkY95WJhZDvF03N8y22YCN/fB+XjpL5f/y5tLCIM/tkBXuxk/6x+m92k9cogHDXKwPSBzoUDoEMNjTe6G/d9l+7rm84rhEe/MwpHTIT0RgNTAwsHiBNg+kC4rY5x/ysL5f5mRs7xFY7P+4rrMJBwgJwzppc4IPSkHxOhCYaXu8BbPQMbup1T4OM+xn35nuDLK2KzySvUvhr0b+hsBIgIg8/ugJSy3n0K+CItdOFQCgZ+7RUOEBWqx0TFgt0w/GZY9ag4Y/Rk5Yg2oCcqHF6+BdYMlxGk/xTjCFzgEfvp2832dfLb9JbtNz7oHfWhZFTghzTTo5Zx+/AZX0NwZrqvcTWhLzSvElxZ4WEwsZ/0KEWczIVP1xjPy8qBL9Yb9z3SQnF/k+DKA/hHWxja3Cjxry+WEc0JA2+Unswp/RtCzXLGfTuzQmuUx0wdRLUywV2vaWKb6DkXhHPEzIFsWHnA91k8SqP/FMg4A6lJsHQoLB4C96UqKsfL6LjigEZMBDRJgvd7w8GR8HxHEZRRi8TBZMX0AALit688fNq43bSy+MKDxUr/z8qBCnHyt1LildHTvjrc1TDoogAZSXrWNgrA0n2i2hXx1lI4rxsdK8TC2N6hNTJNg3/31Ph2syIrR+6RWyBzECPb2V8bFQ6jugZf5h31YPRS7/YZPw6NQFQ12YI/bYfsXILSEv7ZEUPHEqyQ6QnXkCZmofSfOA9P/ART7hKBaFcd2lXXUMo7D6dpMrLp2XsSRi/2X2agr+5XQKYPNKomCXEB7mTCo2TO5NVFvsf0Kla+B1aahsf7GxfP0GuYYFTLzucbj68wlTeoicxWh0pcFPStrzFONzH3807FyHb2D9GtJiSXDr68GmWN26G6MFMrGbczz4mN8809zoWkUUX5XQoqlxI1b9Ee6+PTN0uD1wuhpvl3woA4b/SdoZlBAbQGvwKSFG/eY/zYSom6dOysVGDtIXnBGzIUW49rHMz2+qft2JapOJfvvXd0ONybGvg6O55qJz8rLhTIbLyeuxsVr7yie+gFZPVB6d3sBL1JUmijcvWygc9xwvWl4cEmMEFnaM/fBVVHi0o9qImoLFYOh8vFtAHQ+VOZyTeTVwifrhXbwgk5+UYngh4NGN1dIjHsCCoWa/l+ke7f9il+2qZZutyC/eCL9hjPr1ravkcoLrtP+E52htKLm2mSJEN/0b3PXJDR0c7wjvI3sRKAEC+z5P3ekq+uN9jP5MHE9fIDGWl61pKe+8aKYhterpis8nGwejg8N0/x7nKNfJODZc525wLy2z7r+bS6FWBcH2hTLfA9AgqIR4kP/p1lsPH//d7O346+0Vhx7KyxF60Q6//cS4F59jkiDMpfgjJjIuC6KJkUBVGjdxy3z+K7nB2BU2KjxC09bKa4QK0+VVqG/IooFQOda0DP2op7UzVHXr9giAqHt3po/L2lOHA+WgW7sqTDWXMIcvKk3oHYqnP8xEXB7XXgrkYSw+U0hs720TLPwYCp4n93QnS4zAOUjoGmlUWFaFpZo/YY/9d4lLGmTly6xcHsQw8rMgyLSZjmW/dgwl2uJqVixPgd1kI6wtnbfN+TnuxcmJ4O09M1npwtXrhhLST27lKSXBr+pxV0rQnzd8L6w7B4rwQiOunTlILuNcUV3buuTFIH44AAGwE5lwfdJsD6DN9jKWWhey2oVloks2mSDMOR4eahVwvo7jSrILk2BtWlwPwR8wtlgbXi9ub5Ht+6X0trUGkadKwhv/P58EUabD4qKvXGI/6vy86Fp+dIeM/MgZBcDC+WHqXgveXwyRr4XWePxERImJETSkbBnB3y99iV0ok90ASGNpeZdyf4FZBn5hqFI0yDQY1lprlewqXTQc0qyP5TBDRui0PdBOO2QsIPimv4nsszOiXCNEhyEJb+R6REpDQikPeTVyAz89uzxMExY7PvXFZaBnSZIBGz/mLynLJ8Pwz6BssA0B61fF25/kg1zdF4FHy2Vgz3h5rCu7dCTIB0DMuicgtgkmki7bWu8FlfadDBNN5A0b+tk40aTuY5Y85EKOQWyAcs+ukNtQpxUCbGWCm7HtIp0343bqeUvXxCfiXRkKiBFlXFuziuD2Q8C0sfliBCPTuy4Pl5xStv3g7o8Im1cIDk/TilcSWoU953v0fB+DXQ7XN8nABmLAVkwS5joypTAh5rHZqqnhkg2jQhToLGilBIYlFxeHspVBjl/Q2a5j0WpkGP2sYnmWjqDEJhlimcv0WQUQDXEhFh4gH6eTDcbZrQ/dmhvWqFUjB0pnVoEEgw50NNnbdCTfPmilixZC+8sdT/cfAjIOaciMrXhZ4QM3+n/fHwMPhLPeO+catFtw2VlQeN2+VNk5zdahq3f9wGvxdjFFm23zeUYVDj0O93pXh8Fgz82vubuyP4e7x0i7Hj/P1I4F7ZH1syjculmnm6nVElOpsHqw/Cm0tg9GLFhLUSrqIfY0a0MsaLmVkXINXAUkDMMnomL7hEqSJy8uH1JYHP04eBgMTcPDU7tIT77cdhrqmx9qplvNF9qcYwi/xCGDzdd8bdCdm50tD0JdQuD51S/F7yh+HHrWJcF/3MKa1OKBllVCUVoS+UUKWUeEKtaJsMI26Wv7dlwpM/QflR0OJDeHouTNmoMXg6JL8BzcYqJqdJcllUuKiF/ty65QJkuloKSENT6MCh09LwgiG/ELpPEAM4EA0rSuCZnvFrJMgsmHedWwD3TYM8nYuy0nXQu67x7WgavGmKMl5zSPJHLgThRcsvhH5TZKJNz6tdnRuSV5MGJgfJTItI60CsOmjsPKuVhqgQjfT4aGuVqEEiTB0g2sZzc6HRe4p//2b8Vi0vRvgqYN1hjfumQcN3xeDvlGKdYhEXqXiqrf0DW37GtsnGiNgCjwzB5uhPfyzbJ2mYS/dZHzdXSUMmhiqbwltemA99v7QfdotIPwptx/k21je6WzfWfg3g9rrGfbO2ScL/iv32ZSkFS/ZA4/d9Vch7GvmqjH9U+pjquWSvrAbidP5m4W7JsdDTvnrx6vTurfBCR0X1MlAlXkb7hUMgLhK6fg6vLYG8QmOHV7OcdeT3lkwx+Ceth6Et4Lt7FU2SxO5tkwwLh2gBc1f8JkyNnONrwMRFSnh3aiVoVlnitTREJVp7SNyAX//u20jNrH4Umlk80OI9klNuVnU0FANTNdpXV9xQVqNGWTHkdp2QVVbm7ZRANvOD9K2vmHaP5te5cDoX2o03ZsYV0e0G6FlH8mCuLy1Cse+UhKrMTJcJKzPNK8OyR/yPHv4WbQiWBbug82fe7VbXy2y4HicJU+fypCPbZtIOysXCiJZQp4K49GuUkdVLjp4VG2P3CQn5mLXdGAURFwkbRyhqlL207juloMfnMNePPTuoMbzTS1HpNc1nNROQ0Jy5D0DnAHFXVvgdDF/sLDPo+nzpc/mSZFLExUxUW/ukRy3pkU/qjO7jfiYP21eH+Q+KkOi9aAqNyWkwOYhknNvqwJf9/QsHSDrpz4Oh71ewdK/x2Nyd/j+IFV1SJAr2WlCtioiLgv/eBe0/Mb7vrBx48Rcnd/C+3TBNpgEutXCAjFR232JocyhTQqNjDe/EoJ5CBf9aEJqA+P2csZGSz922mv+LPfgXjqhweKwVzBjoG52bbhGpWUSrZNjymAhWKK86NhL+1UkxfaCzzMAKcZJH/nTb0Nbwio2ElzrDj/f75m9fC6QmwfJHiheynhQPs++HO0PM4QnEVhv796YqEkIC8GwH5dcYD2a1Gj22/V35OMncGt9HcnudUCJSUiNX/1VSUqMjJG1Wz9u/2meeJcXDrPth3mDoV9+ZoJSIELVg49/hxc7O1qkqIjpCQp/XD4e/3exsJjg+WvLI0x+D/+10edeFutzUS4B1w2WNqZYOQzAA6ieImrj7SUXXmpdv9cdmla33a8CHt3tj4NomawxoZN1jh9oBOF5Z0eORmJi0DDF+snPFDogOl54zpayEETerYt3ANh01xirVLOc8cCzznNg1G4/IioE5+TKkl4ySEOwGidKTOInwdML5/Ish4Edhz0lvCElcpCQr1U8UozDYkIqtmUYDuHK8s1USzZy+YPQqlowSe0FPTr6v29ZfQ9OjlEQypGXI8x87K/dSSCdUIQ7qJUKDBKhZ/sotx/rwDInLKiJck9UShzQ3nncqV7yn+rmw+GhY+BA0cfD8Ztz/D+JyTeBR8GUazEgXLWVIM1nm1orcAkmrXnFAOqGR7XyzMJ3iCoiLiw3XkM/FxeXK4wqIi4sNroC4uNjgCoiLiw2ugLi42OAKiIuLDa6AuLjY4AqIi4sNroC4uNjgCoiLiw2ugLi42OAKiIuLDa6AuLjY8H95fXySW2Mr8wAAAABJRU5ErkJggg=="

# ─── Styles ───────────────────────────────────────────────────────────────────
DATA_FILL = PatternFill("solid", fgColor="C5E1FF")
WRAP_TOP  = Alignment(wrap_text=True, vertical="top")
DATE_FMT  = '[$-409]mmm dd\\, yyyy;@'

def seg(bold=False, italic=False, color="000000", size=10):
    return Font(name="Segoe UI Light", size=size, bold=bold, italic=italic, color=color)

def fr(formula, fill_hex=None, bold=False, italic=False, font_color="000000"):
    fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None
    font = Font(name="Segoe UI Light", size=10, bold=bold, italic=italic, color=font_color)
    return FormulaRule(formula=[formula], fill=fill, font=font)

def cidx(col): return column_index_from_string(col)

COLS = [
    ("A","sds_id"), ("B","product_name"), ("C","supplier"), ("D","sds_date"),
    ("E","composition"),
    ("F","cytotoxicity"), ("G","sensitisation"), ("H","skin_irritation"),
    ("I","eye_irritation"), ("J","acute_systemic_toxicity"),
    ("K","subacute_subchronic"), ("L","chronic_toxicity"),
    ("M","genotoxicity"), ("N","carcinogenicity"), ("O","reproductive_toxicity"),
    ("P","endocrine_disruption"), ("Q","bioaccumulation"),
    ("R","haemocompatibility"), ("S","pyrogenicity"), ("T","implantation"),
    ("U","immune_responses"), ("V","other_biological_effects"),
    ("W","cmr_regulatory_status"), ("X","azo_dyes"),
    ("Y","formaldehyde"), ("Z","heavy_metals"), ("AA","reach_svhc"),
    ("AC","alert_justification"),
]

ANALYSIS_FIELDS = [
    "cytotoxicity","sensitisation","skin_irritation","eye_irritation",
    "acute_systemic_toxicity","subacute_subchronic","chronic_toxicity",
    "genotoxicity","carcinogenicity","reproductive_toxicity",
    "endocrine_disruption","bioaccumulation","haemocompatibility",
    "pyrogenicity","implantation","immune_responses",
    "other_biological_effects","cmr_regulatory_status","azo_dyes",
    "formaldehyde","heavy_metals","reach_svhc",
]

ALERT_COLORS = {
    "CRITICAL":   ("#FF0000", "#FFFFFF", "🔴"),
    "MAJOR":      ("#FF8C00", "#FFFFFF", "🟠"),
    "MINOR":      ("#FFC000", "#000000", "🟡"),
    "NONE":       ("#00B050", "#FFFFFF", "🟢"),
    "NOT AN SDS": ("#808080", "#FFFFFF", "⬜"),
    "ERROR":      ("#888888", "#FFFFFF", "⚠️"),
}

DB_DISPLAY_NAMES = [
    ("CLP Annex VI ATP22",          "CLP Annex VI ATP22"),
    ("SVHC Candidates",             "SVHC Candidates (ECHA)"),
    ("REACH Annex XVII",            "REACH Annex XVII"),
    ("Aromatic Amines (Ent.43)",    "Aromatic Amines (Entry 43)"),
    ("Azo - Restricted Amines (JTF)", "Azo Restricted Amines (JTF)"),
    ("IARC",                        "IARC Monographs"),
    ("ED Assessment (ECHA)",        "ED Assessment (ECHA)"),
    ("TEDX",                        "TEDX"),
]

# ─── SDS format detection ─────────────────────────────────────────────────────
def detect_sds_format(text: str) -> str:
    """Detect SDS format — works for EN/FR/DE/ES/IT documents."""
    high_sections = 0
    for i in range(9, 17):
        patterns = [
            rf'section\s+{i}\b',       # EN
            rf'rubrique\s+{i}\b',      # FR
            rf'abschnitt\s+{i}\b',     # DE
            rf'secci[o\u00f3]n\s+{i}\b', # ES
            rf'sezione\s+{i}\b',       # IT
            rf'^\s*{i}[\s\.\:\-\)]', # Generic numbered
        ]
        if any(re.search(p, text, re.IGNORECASE | re.MULTILINE) for p in patterns):
            high_sections += 1
    return "16-section" if high_sections >= 4 else "8-section"

MSDS_8_MODIFIER = """

## IMPORTANT — OLD FORMAT MSDS DETECTED (pre-GHS / pre-2015)

This document appears to be an old-format MSDS (pre-GHS harmonisation, typically 8–16 sections
with non-standardised structure). Apply the following rules:

- Section mapping: hazard info may appear in "Section II/III", "Health Hazards", "Chemical Hazards"
  instead of current Sections 2 and 11.
- CLP H-phrases may be absent; old R-phrases may be used — map R-phrases to H-phrases where possible.
- Composition may be in "Section II" or "Ingredients" — extract all CAS numbers found.
- When a specific field CANNOT be found due to the old format, write EXACTLY:
  "Not available in MSDS format — [field name] section absent or non-standardised"
- Do NOT invent, extrapolate, or assume data not explicitly present in the document.
- Still output the complete JSON block with all required keys — use the above phrase for unavailable fields.

"""

# ─── Database helpers ─────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def get_db_dates(db_path: str) -> dict:
    """Read version/date from each database sheet (cell A4, fallback first 3 rows)."""
    if not os.path.exists(db_path):
        return {}
    dates = {}
    try:
        wb = load_workbook(db_path, read_only=True, data_only=True)
        for sheet_name, _ in DB_DISPLAY_NAMES:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            val = None
            try:
                val = ws["A4"].value
            except Exception:
                pass
            if not val or str(val).strip() in ("", "None"):
                for row in ws.iter_rows(max_row=3, values_only=True):
                    for cell in row:
                        if cell and str(cell).strip() not in ("", "None"):
                            val = str(cell).strip()
                            break
                    if val:
                        break
            dates[sheet_name] = str(val).strip() if val else "—"
        wb.close()
    except Exception:
        pass
    return dates


@st.cache_data(show_spinner=False)
def load_databases(db_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(db_bytes), read_only=True, data_only=True)
    dbs = {}

    def find_header_row(ws, keyword, max_rows=10):
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), 1):
            if any(keyword.lower() in str(c).lower() for c in row if c):
                return i
        return 1

    def sheet_to_df(name, cas_hint="CAS"):
        if name not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb[name]
        h = find_header_row(ws, cas_hint)
        rows = list(ws.iter_rows(min_row=h, values_only=True))
        if not rows:
            return pd.DataFrame()
        hdrs = [str(c).strip() if c else f"_c{i}" for i, c in enumerate(rows[0])]
        return pd.DataFrame(rows[1:], columns=hdrs)

    def register(key, sheet_name, cas_hint="CAS"):
        try:
            df = sheet_to_df(sheet_name, cas_hint)
            cas_col = next((c for c in df.columns if "cas" in c.lower()), None)
            if cas_col is not None:
                df[cas_col] = df[cas_col].astype(str).str.strip()
                dbs[key]            = df
                dbs[f"{key}_cas"]   = cas_col
                dbs[f"{key}_sheet"] = sheet_name
        except Exception:
            pass

    register("clp",     "CLP Annex VI ATP22")
    register("svhc",    "SVHC Candidates")
    register("reach",   "REACH Annex XVII")
    register("amines",  "Aromatic Amines (Ent.43)")
    register("azo_jtf", "Azo - Restricted Amines (JTF)")
    register("iarc",    "IARC")
    register("ed",      "ED Assessment (ECHA)")
    register("tedx",    "TEDX")
    wb.close()
    return dbs


def lookup_all_cas(cas_list, dbs):
    if not cas_list:
        return "No CAS numbers extracted from the SDS text."
    today = datetime.date.today().strftime("%b %Y")
    lines = [f"PRE-COMPUTED DATABASE LOOKUP — {len(cas_list)} CAS extracted", "="*60]
    for cas in cas_list:
        hits = []
        if "clp" in dbs:
            cc = dbs["clp_cas"]
            m  = dbs["clp"][dbs["clp"][cc].str.contains(cas, regex=False, na=False)]
            for _, r in m.iterrows():
                codes = str(r.get("Hazard Statement Code(s)","")).replace("\n"," ")
                cat   = str(r.get("Hazard Class and Category Code(s)",""))
                name  = r.get("International Chemical Identification","")
                hits.append(f"  [CLP Annex VI ATP22] {name} | {cat} | H-codes: {codes}")
        if "svhc" in dbs:
            cc = dbs["svhc_cas"]
            m  = dbs["svhc"][dbs["svhc"][cc].str.contains(cas, regex=False, na=False)]
            for _, r in m.iterrows():
                reason = r.get("Reason for inclusion", r.get("reason",""))
                name   = r.get("Substance name", r.get("name",""))
                hits.append(f"  [SVHC Candidates] {name} — Reason: {reason}")
        if "iarc" in dbs:
            cc = dbs["iarc_cas"]
            m  = dbs["iarc"][dbs["iarc"][cc].str.contains(cas, regex=False, na=False)]
            for _, r in m.iterrows():
                grp   = r.get("Group", r.get("group",""))
                agent = r.get("Agent", r.get("agent",""))
                hits.append(f"  [IARC, {today}] Group {grp} — {agent}")
        if "reach" in dbs:
            cc = dbs["reach_cas"]
            m  = dbs["reach"][dbs["reach"][cc].str.contains(cas, regex=False, na=False)]
            for _, r in m.iterrows():
                entry = r.get("Entry", r.get("entry",""))
                restr = str(r.get("Restriction", r.get("restriction","")))[:120]
                hits.append(f"  [REACH Annex XVII] Entry {entry} — {restr}")
        if "amines" in dbs:
            cc = dbs["amines_cas"]
            m  = dbs["amines"][dbs["amines"][cc].str.contains(cas, regex=False, na=False)]
            if not m.empty:
                hits.append("  [Aromatic Amines Ent.43] MATCH — restricted aromatic amine")
        if "azo_jtf" in dbs:
            cc = dbs["azo_jtf_cas"]
            m  = dbs["azo_jtf"][dbs["azo_jtf"][cc].str.contains(cas, regex=False, na=False)]
            if not m.empty:
                hits.append("  [Azo - Restricted Amines JTF] MATCH — restricted azo amine")
        if "ed" in dbs:
            cc = dbs["ed_cas"]
            m  = dbs["ed"][dbs["ed"][cc].str.contains(cas, regex=False, na=False)]
            for _, r in m.iterrows():
                status = str(r.get("Status", r.get("status","")))
                hits.append(f"  [ED Assessment (ECHA), {today}] {status}")
        if "tedx" in dbs:
            cc = dbs["tedx_cas"]
            m  = dbs["tedx"][dbs["tedx"][cc].str.contains(cas, regex=False, na=False)]
            if not m.empty:
                hits.append("  [TEDX] Listed as endocrine disruptor")
        if hits:
            lines.append(f"\nCAS {cas}:")
            lines.extend(hits)
        else:
            lines.append(f"\nCAS {cas}: Not found in CLP ATP22, SVHC, IARC, REACH Annex XVII, "
                         "Aromatic Amines, ED Assessment, TEDX")
    return "\n".join(lines)


# ─── Excel helpers ────────────────────────────────────────────────────────────
def parse_date(s):
    if not s or str(s).strip() in ("","Not stated","N/A","not stated"):
        return str(s) if s else "Not stated"
    for fmt in ("%d/%m/%Y","%m/%d/%Y","%Y-%m-%d","%d-%m-%Y",
                "%B %d, %Y","%b %d, %Y","%d %B %Y","%d %b %Y"):
        try:
            return datetime.datetime.strptime(str(s).strip(), fmt).strftime("%b %d, %Y")
        except Exception:
            pass
    return str(s)


def compute_alert_level(data):
    all_text = " ".join(str(data.get(f,"")) for f in ANALYSIS_FIELDS)
    if "🔴" in all_text: return "CRITICAL"
    if "🟠" in all_text: return "MAJOR"
    if "🟡" in all_text: return "MINOR"
    return "NONE"


def apply_cf(ws):
    ws.conditional_formatting.add("AB3:AB9999", fr('$AB3="CRITICAL"',   "FF0000", True,  False, "FFFFFF"))
    ws.conditional_formatting.add("AB3:AB9999", fr('$AB3="MAJOR"',      "FF8C00", True,  False, "FFFFFF"))
    ws.conditional_formatting.add("AB3:AB9999", fr('$AB3="MINOR"',      "FFC000", True,  False, "000000"))
    ws.conditional_formatting.add("AB3:AB9999", fr('$AB3="NONE"',       "00B050", False, False, "FFFFFF"))
    ws.conditional_formatting.add("AB3:AB9999", fr('$AB3="NOT AN SDS"', "808080", False, False, "FFFFFF"))
    ws.conditional_formatting.add("A3:AH9999",  fr('$AB3="CRITICAL"', "FFE6E6"))
    ws.conditional_formatting.add("A3:AH9999",  fr('$AB3="MAJOR"',    "FFF2E6"))
    ws.conditional_formatting.add("A3:AH9999",  fr('$AB3="MINOR"',    "FFFDE6"))
    ws.conditional_formatting.add("AG3:AG9999",
        fr('$AG3="AI - SDS Screening agent only"', "FFF2CC", False, True))
    ws.conditional_formatting.add("AG3:AG9999",
        fr('AND($AG3<>"",$AG3<>"AI - SDS Screening agent only")', "E2EFDA"))
    ws.conditional_formatting.add("AD3:AD9999", fr('$AD3<>""', "E2EFDA", True))
    ws.conditional_formatting.add("AH3:AH9999", fr('$AH3="Draft"',      "BDD7EE"))
    ws.conditional_formatting.add("AH3:AH9999", fr('$AH3="Validated"',  "00B050", True, False, "FFFFFF"))
    ws.conditional_formatting.add("AH3:AH9999", fr('$AH3="Superseded"', "D9D9D9", False, False, "808080"))
    ws.conditional_formatting.add("AH3:AH9999", fr('$AH3="Archived"',   "595959", False, False, "FFFFFF"))
    ws.conditional_formatting.add("D3:D9999",
        fr('AND(D3<>"",ISNUMBER(D3),D3<TODAY()-1095)', "FFE0E0", False, False, "CC0000"))
    ws.conditional_formatting.add("E3:AC9999", FormulaRule(
        formula=['NOT(ISERROR(SEARCH("TO VERIFY",E3)))'],
        fill=PatternFill("solid", fgColor="E2CCFF"),
        font=Font(name="Segoe UI Light", size=10),
    ))


def write_row(wb, data, row):
    ws = wb["SDS Analysis"]
    for col_letter, key in COLS:
        val = data.get(key,"")
        if key == "sds_date":
            val = parse_date(str(val)) if val else "Not stated"
        c = ws.cell(row=row, column=cidx(col_letter), value=val)
        c.fill = DATA_FILL; c.alignment = WRAP_TOP; c.font = seg()

    ab = ws.cell(row=row, column=cidx("AB"))
    ab.value = (f'=IF(COUNTIF(F{row}:AA{row},"*🔴*")>0,"CRITICAL",'
                f'IF(COUNTIF(F{row}:AA{row},"*🟠*")>0,"MAJOR",'
                f'IF(COUNTIF(F{row}:AA{row},"*🟡*")>0,"MINOR","NONE")))')
    ab.fill = DATA_FILL; ab.alignment = Alignment(horizontal="center", vertical="center"); ab.font = seg(bold=True)

    ws.cell(row=row, column=cidx("AD"), value="").fill = DATA_FILL
    ws.cell(row=row, column=cidx("AE"), value="").fill = DATA_FILL

    today = datetime.date.today()
    af = ws.cell(row=row, column=cidx("AF"),
                 value=datetime.datetime(today.year, today.month, today.day))
    af.fill = DATA_FILL; af.alignment = WRAP_TOP; af.font = seg(); af.number_format = DATE_FMT

    ag = ws.cell(row=row, column=cidx("AG"), value="AI - SDS Screening agent only")
    ag.fill = DATA_FILL; ag.alignment = WRAP_TOP; ag.font = seg(italic=True)

    ah = ws.cell(row=row, column=cidx("AH"), value="Draft")
    ah.fill = DATA_FILL; ah.alignment = WRAP_TOP; ah.font = seg()


def extend_dashboard(wb, n):
    if "SDS Dashboard" not in wb.sheetnames:
        return
    ws = wb["SDS Dashboard"]
    if ws.cell(row=n, column=1).value is not None:
        return
    formulas = [
        ("A", f"=ROW()-2"),
        ("B", f"='SDS Analysis'!$A{n}"),
        ("C", f"='SDS Analysis'!$B{n}"),
        ("D", f"='SDS Analysis'!$C{n}"),
        ("E", f"='SDS Analysis'!$D{n}"),
        ("F", f"='SDS Analysis'!$AE{n}"),
        ("G", f"='SDS Analysis'!$AC{n}&IF('SDS Analysis'!$AD{n}<>\"\",CHAR(10)&CHAR(10)&\"Toxicologist comment: \"&'SDS Analysis'!$AD{n},\"\")" ),
        ("H", f"='SDS Analysis'!$AF{n}"),
        ("I", f"='SDS Analysis'!$AG{n}"),
        ("J", f"='SDS Analysis'!$AH{n}"),
    ]
    for col_l, val in formulas:
        c = ws.cell(row=n, column=cidx(col_l), value=val)
        c.fill = DATA_FILL; c.alignment = WRAP_TOP; c.font = seg()
        if col_l in ("E","H"):
            c.number_format = DATE_FMT


def write_project_tab(wb, config, results):
    if "Project" not in wb.sheetnames:
        wb.create_sheet("Project", 0)
    ws = wb["Project"]
    ws.delete_rows(1, ws.max_row + 1)
    blue  = PatternFill("solid", fgColor="007AFF")
    light = PatternFill("solid", fgColor="EAF4FF")

    def row_w(r, a, b="", a_fill=None, a_font=None):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ca.fill = a_fill or PatternFill()
        ca.font = a_font or seg()
        cb.font = seg()
        ca.alignment = Alignment(vertical="center", wrap_text=True)
        cb.alignment = Alignment(vertical="center", wrap_text=True)

    row_w(1, "ACONIS BIOCOMPATIBILITY SCREENING — SESSION RECORD",
          a_fill=blue, a_font=seg(bold=True, color="FFFFFF", size=13))
    row_w(2, "")
    row_w(3, "Generated:",    datetime.datetime.now().strftime("%b %d, %Y — %H:%M"))
    row_w(4, "Tool Version:", "v2.0 — Claude Sonnet 4.5 | ISO 10993-1:2025 / MDR 2017/745")
    row_w(5, "")
    row_w(6, "CLIENT", a_fill=blue, a_font=seg(bold=True, color="FFFFFF", size=11))
    row_w(7, "Client Name:",  config.get("client_name",""))
    row_w(8, "Client Email:", config.get("client_email",""))
    row_w(9, "")
    row_w(10, "ANALYSIS SUMMARY", a_fill=blue, a_font=seg(bold=True, color="FFFFFF", size=11))
    row_w(11, "Files Processed:", str(len(results)))
    row_w(12, "")

    hdrs = ["#","SDS ID","Product Name","Supplier","SDS Date","Alert Level","Alert Justification (summary)","Status"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=13, column=i, value=h)
        c.fill = blue; c.font = seg(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(results, 1):
        justif = r.get("alert_justification","")
        tox_comment = r.get("tox_comment","")
        if tox_comment:
            justif += f"\n\nToxicologist comment: {tox_comment}"
        row_data = [i, r.get("sds_id",""), r.get("product",""),
                    r.get("supplier",""), r.get("sds_date",""),
                    r.get("alert_level",""), justif,
                    r.get("status","Draft")]
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=13+i, column=j, value=val)
            c.fill = light; c.font = seg()
            c.alignment = Alignment(vertical="top", wrap_text=True)

    for col, w in zip("ABCDEFGH", [6,22,35,25,14,14,60,12]):
        ws.column_dimensions[col].width = w


# ─── Prompt builder ───────────────────────────────────────────────────────────
def build_system_prompt(base_prompt, config, sds_format="16-section"):
    azo_instr = (
        "Col X (Azo dyes / Aromatic amines): write exactly "
        "`N/A — not applicable: non-textile device [Agent assessment]`"
    )
    prompt = base_prompt
    prompt = prompt.replace("<<CLIENT_NAME>>",          config.get("client_name","CLIENT"))
    prompt = prompt.replace("<<DEVICE_DESCRIPTION>>",   config.get("device_description","medical device"))
    prompt = prompt.replace("<<CONTACT_SCENARIO>>",     "General screening — contact scenario not specified")
    prompt = prompt.replace("<<AZO_DYES_INSTRUCTION>>", azo_instr)

    if sds_format == "8-section":
        prompt += MSDS_8_MODIFIER

    json_schema = """\n\n---\n\n## PYTHON APP — MANDATORY JSON OUTPUT\n
After your screening report, output a JSON block in ```json ... ```.
Use EXACTLY these keys. Values must include all markers (🔴🟠🟡), source tags, newlines as \\n.\n
```json
{
  "sds_id": "", "product_name": "", "supplier": "", "sds_date": "", "composition": "",
  "cytotoxicity": "", "sensitisation": "", "skin_irritation": "", "eye_irritation": "",
  "acute_systemic_toxicity": "", "subacute_subchronic": "", "chronic_toxicity": "",
  "genotoxicity": "", "carcinogenicity": "", "reproductive_toxicity": "",
  "endocrine_disruption": "", "bioaccumulation": "", "haemocompatibility": "",
  "pyrogenicity": "", "implantation": "", "immune_responses": "",
  "other_biological_effects": "", "cmr_regulatory_status": "", "azo_dyes": "",
  "formaldehyde": "", "heavy_metals": "", "reach_svhc": "", "alert_justification": ""
}
```
"""
    return prompt + json_schema


# ─── Email ─────────────────────────────────────────────────────────────────────
def send_email_with_attachment(to_addr, subject, body_html,
                                attachment_bytes=None, attachment_name=None):
    smtp_user = st.secrets.get("SMTP_USER","")
    smtp_pass = st.secrets.get("SMTP_PASSWORD","")
    if not smtp_user or not smtp_pass:
        st.warning("⚠️ Email not sent — SMTP credentials missing in Secrets.")
        return False
    try:
        msg = MIMEMultipart("mixed")
        msg["From"] = smtp_user; msg["To"] = to_addr; msg["Subject"] = subject
        msg.attach(MIMEText(body_html, "html", "utf-8"))
        if attachment_bytes and attachment_name:
            part = MIMEBase("application","octet-stream")
            part.set_payload(attachment_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{attachment_name}"')
            msg.attach(part)
        with smtplib.SMTP("smtp.office365.com", 587) as server:
            server.starttls(); server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_addr, msg.as_string())
        return True
    except Exception as e:
        st.error(f"❌ Email error: {e}"); return False


# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] { font-family: 'Segoe UI Light', 'Helvetica Neue', sans-serif; }
[data-testid="stSidebar"] { background-color: #F0F6FF; }
[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,[data-testid="stSidebar"] h3 { color: #007AFF; }
h1 { color: #007AFF !important; font-weight: 300; }
h2, h3 { color: #007AFF; font-weight: 300; }
.stButton > button[kind="primary"] {
    background-color: #007AFF !important; border: none;
    font-family: 'Segoe UI Light', sans-serif; letter-spacing: 0.03em;
}
.stButton > button[kind="primary"]:hover { background-color: #005FCC !important; }
.footer-aconis {
    margin-top: 40px; padding: 18px 0 8px 0;
    border-top: 1px solid #DCE8FF; color: #888;
    font-size: 0.78em; text-align: center; line-height: 1.7;
}
</style>
""", unsafe_allow_html=True)

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    logo_bytes = base64.b64decode(LOGO_B64)
    st.image(logo_bytes, use_container_width=True)
    st.markdown("---")
    st.caption("v2.0 — Claude Sonnet 4.5 — ISO 10993-1:2025")
    st.markdown("---")
    st.markdown(
        '<a href="mailto:elodie.saudrais@aconis.fr?subject=SDS%20Screener%20-%20Contact" '
        'style="display:block;text-align:center;background:#007AFF;color:#fff;'
        'padding:8px 0;border-radius:6px;text-decoration:none;font-size:0.9em;">'
        '✉️ Contacter ACONIS</a>',
        unsafe_allow_html=True,
    )

api_key     = st.secrets.get("ANTHROPIC_API_KEY","")
ADMIN_EMAIL = st.secrets.get("ADMIN_EMAIL","elodie.saudrais@aconis.fr")
BETA_LIMIT  = 5

# ─── Main area ────────────────────────────────────────────────────────────────
st.title("🔬 SDS Biocompatibility Screener")
st.caption("ACONIS — ISO 10993-1:2025 / MDR 2017/745 — Claude AI")

# ── 1. Presentation ───────────────────────────────────────────────────────────
db_dates = get_db_dates(DB_FILE) if os.path.exists(DB_FILE) else {}
db_rows_html = "".join(
    f'<tr style="border-bottom:1px solid #DCE8FF;">'
    f'<td style="padding:8px 14px;color:#222;font-weight:500;">{disp}</td>'
    f'<td style="padding:8px 14px;color:#555;font-size:0.88em;">Last update: {db_dates.get(sheet, "—")}</td></tr>'
    for sheet, disp in DB_DISPLAY_NAMES
)

st.markdown(f"""
<div style="background:#F0F6FF;border-left:4px solid #007AFF;border-radius:8px;
            padding:22px 26px;margin:10px 0 24px 0;">

<p style="margin:0 0 12px 0;font-size:1.05em;color:#007AFF;font-weight:600;">About this application</p>

<p style="margin:0 0 12px 0;color:#222;line-height:1.7;">
This application automatically analyses <b>Safety Data Sheets (SDS)</b> to support the biocompatibility
assessment of medical devices (ISO&nbsp;10993-1:2025 / MDR 2017/745).<br>
It goes beyond simply reading the SDS: it <b>identifies every chemical substance</b> listed in the document
and <b>automatically cross-references each one against 8 regulatory databases</b> — with no AI at this stage.
Only then does an AI model perform an expert-level synthesis, strictly constrained to document content
(no invention, uncertain items flagged as "TO VERIFY").
</p>

<div style="background:#fff;border:1px solid #DCE8FF;border-radius:8px;
            padding:14px 20px;margin:0 0 16px 0;font-size:0.87em;color:#333;line-height:1.8;">
  <b style="color:#007AFF;">How it works:</b><br>
  📄 <b>SDS PDF</b>
  &nbsp;→&nbsp; 🔢 <b>Automated text &amp; CAS extraction</b> <span style="color:#888;">(rule-based, no AI)</span>
  &nbsp;→&nbsp; 🔍 <b>Automated cross-check — 8 databases</b> <span style="color:#888;">(rule-based, no AI)</span>
  &nbsp;→&nbsp; 🤖 <b>AI-assisted synthesis</b> <span style="color:#888;">(Claude AI — document content only)</span>
  &nbsp;→&nbsp; 📊 <b>Excel report + email</b>
</div>

<table style="border-collapse:collapse;width:100%;margin:0 0 14px 0;
              border:1px solid #DCE8FF;border-radius:6px;overflow:hidden;">
<tr style="background:#007AFF;color:#fff;">
  <th style="padding:8px 14px;text-align:left;font-weight:500;">Database</th>
  <th style="padding:8px 14px;text-align:left;font-weight:500;">Version</th>
</tr>
{db_rows_html}
</table>

<div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:12px;">
  <span style="background:#EAF4FF;border-radius:6px;padding:7px 16px;color:#007AFF;font-size:0.87em;">
    ⏱ ~1–2 min per SDS
  </span>
  <span style="background:#EAF4FF;border-radius:6px;padding:7px 16px;color:#007AFF;font-size:0.87em;">
    🔒 No data stored on server — transmitted by email only
  </span>
</div>

<hr style="border:none;border-top:1px solid #DCE8FF;margin:10px 0;">
<p style="margin:0;color:#888;font-size:0.84em;">
  ⚠️ <i>Application under development — feedback welcome:</i>
  <a href="mailto:elodie.saudrais@aconis.fr?subject=SDS Screener - Suggestion"
     style="color:#007AFF;">elodie.saudrais@aconis.fr</a>
  &nbsp;·&nbsp;
  🧪 <b>Beta</b> — limited to <b>5 SDS per session</b>.
</p>
</div>
""", unsafe_allow_html=True)

# ── 2. Identification ─────────────────────────────────────────────────────────
st.subheader("1 — Identification")
col1, col2 = st.columns(2)
with col1:
    client_name  = st.text_input("Company name", placeholder="e.g. Acme Medical")
with col2:
    client_email = st.text_input(
        "Report recipient email",
        placeholder="contact@client.com",
        help="The Excel report will be sent to this address. Enter your email to receive your SDS analyses.")

config = {"client_name": client_name, "client_email": client_email}

st.divider()

# ── 3. SDS Upload ─────────────────────────────────────────────────────────────
st.subheader("2 — SDS Upload")
st.caption("Max 5 files per session (beta)")
pdf_uploads = st.file_uploader(
    "Upload one or more SDS files (PDF)",
    type=["pdf"], accept_multiple_files=True)

st.divider()

# ── 4. Analysis ───────────────────────────────────────────────────────────────
st.subheader("3 — Analysis")
st.caption("Each SDS is processed in ~1–2 min. CAS numbers are extracted and cross-referenced against 8 databases, then analysed by AI.")
can_run = bool(api_key and pdf_uploads and client_email)
run = st.button("🚀 Run Analysis", disabled=not can_run, type="primary", use_container_width=True)
if not can_run and (pdf_uploads is not None or client_email):
    missing = [x for x, ok in [("Anthropic API key", api_key), ("SDS PDF(s)", pdf_uploads), ("Report recipient email", client_email)] if not ok]
    if missing:
        st.info(f"Missing: {', '.join(missing)}")

# ── 5. Processing ─────────────────────────────────────────────────────────────
if run:
    # ── Beta limit check ─────────────────────────────────────────────────────
    is_admin     = client_email.strip().lower() == ADMIN_EMAIL.strip().lower()
    max_files    = None if is_admin else BETA_LIMIT
    if not is_admin and len(pdf_uploads) > BETA_LIMIT:
        st.warning(
            f"⚠️ Beta version: analysis limited to **{BETA_LIMIT} SDS per session**. "
            f"{len(pdf_uploads) - BETA_LIMIT} additional file(s) ignored. "
            
            
        )
        pdf_uploads = pdf_uploads[:BETA_LIMIT]

    missing_files = [(p,n) for p,n in [
        (PROMPT_FILE,"agent_prompt_generic.txt"),
        (DB_FILE,    "databases.xlsx"),
        (TMPL_FILE,  "template.xlsx"),
    ] if not os.path.exists(p)]
    if missing_files:
        for _,n in missing_files:
            st.error(f"❌ Fichier manquant : **{n}**")
        st.stop()

    with open(PROMPT_FILE,"r",encoding="utf-8") as f:
        base_prompt = f.read()

    with st.spinner("Loading databases…"):
        with open(DB_FILE,"rb") as f:
            db_bytes = f.read()
        dbs = load_databases(db_bytes)
    st.success(f"✅ {len([k for k in dbs if k.endswith('_sheet')])} databases loaded")

    with open(TMPL_FILE,"rb") as f:
        tmpl_bytes = f.read()
    wb = load_workbook(io.BytesIO(tmpl_bytes))
    if "SDS Analysis" in wb.sheetnames:
        apply_cf(wb["SDS Analysis"])

    client_api = anthropic.Anthropic(api_key=api_key)
    results    = []

    for idx, pdf_file in enumerate(pdf_uploads):
        st.subheader(f"📄 {pdf_file.name}  ({idx+1}/{len(pdf_uploads)})")
        prog = st.progress(0)
        msg  = st.empty()

        try:
            msg.info("📖 Extracting PDF text…")
            prog.progress(10)
            pdf_bytes = pdf_file.read()
            doc   = fitz.open(stream=pdf_bytes, filetype="pdf")
            pages = [f"--- PAGE {i+1} ---\n{p.get_text()}" for i,p in enumerate(doc)]
            doc.close()
            pdf_text = "\n".join(pages)

            if len(pdf_text.strip()) < 200:
                st.warning("⚠️ Very short text — PDF may be image-based (scanned).")

            sds_format = detect_sds_format(pdf_text)
            fmt_label  = "16-section GHS" if sds_format == "16-section" else "Old MSDS format (8 sections)"

            msg.info("🔢 Extracting CAS numbers…")
            prog.progress(20)
            cas_list = sorted(set(re.findall(r'\b\d{1,7}-\d{2}-\d\b', pdf_text)))
            with st.expander("Technical details"):
                st.caption(f"📋 Format detected: {fmt_label}")
                st.caption(f"{len(cas_list)} CAS found: "
                           f"{', '.join(cas_list[:12])}{'…' if len(cas_list)>12 else ''}") 

            msg.info("📊 Cross-referencing databases…")
            prog.progress(35)
            lookup_report = lookup_all_cas(cas_list, dbs)

            msg.info("🤖 AI analysis in progress (~1–2 min)…")
            prog.progress(50)

            system_prompt = build_system_prompt(base_prompt, config, sds_format)

            config_block = (
                f"DEVICE CONFIGURATION FOR THIS ANALYSIS:\n"
                f"- Client: {config['client_name']}\n"
                f"- SDS Format: {fmt_label}\n"
                f"- Contact Scenario: General screening\n"
            )
            user_msg = (
                f"{config_block}\n"
                f"=== SDS TEXT ===\n{pdf_text[:60000]}\n\n"
                f"=== PRE-COMPUTED DATABASE LOOKUP RESULTS ===\n{lookup_report}\n\n"
                "Use the database results above as authoritative sources. "
                "Complete the full analysis then output the JSON block."
            )

            response = client_api.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=16000,
                system=system_prompt,
                messages=[{"role":"user","content":user_msg}],
            )
            raw = response.content[0].text
            prog.progress(80)

            msg.info("📋 Parsing JSON response…")
            data = {}
            for pattern in [
                r'```json\s*(.*?)\s*```',
                r'```\s*(\{.*?\})\s*```',
                r'\{.*\}',
            ]:
                m = re.search(pattern, raw, re.DOTALL)
                if m:
                    try:
                        data = json.loads(m.group(1) if m.lastindex else m.group())
                        break
                    except Exception:
                        pass

            if not data:
                error_reason = "JSON non trouvé dans la réponse de l'IA · JSON not found in AI response"
                st.warning(f"⚠️ {error_reason}")
                st.text_area("Raw AI response", raw, height=300)
                results.append({
                    "sds_id": "—", "product": pdf_file.name, "supplier": "—",
                    "sds_date": "—", "alert_level": "ERROR",
                    "alert_justification": f"⚠️ Traitement échoué : {error_reason}",
                    "tox_comment": "", "status": "Error", "format": fmt_label,
                })
                prog.progress(100); msg.empty()
                continue

            ws_sds   = wb["SDS Analysis"]
            next_row = 3
            while ws_sds.cell(row=next_row, column=1).value is not None:
                next_row += 1

            msg.info("📝 Writing to Excel…")
            prog.progress(90)
            write_row(wb, data, next_row)
            extend_dashboard(wb, next_row)

            alert_level = compute_alert_level(data)
            results.append({
                "sds_id":             data.get("sds_id",""),
                "product":            data.get("product_name", pdf_file.name),
                "supplier":           data.get("supplier",""),
                "sds_date":           data.get("sds_date",""),
                "alert_level":        alert_level,
                "alert_justification": data.get("alert_justification",""),
                "tox_comment":        "",
                "status":             "Draft",
                "format":             fmt_label,
            })

            prog.progress(100); msg.empty()
            st.success(f"✅ **{data.get('product_name', pdf_file.name)}** — row {next_row} written.")
            with st.expander("📊 Full screening report"):
                st.markdown(raw[:6000] + ("…" if len(raw)>6000 else ""))

        except anthropic.AuthenticationError:
            st.error("❌ Invalid API key."); break
        except anthropic.RateLimitError:
            st.error("❌ Rate limit reached — wait 30 s."); break
        except Exception as exc:
            error_reason = str(exc)
            st.error(f"❌ Error: {exc}")
            results.append({
                "sds_id":"—","product":pdf_file.name,"supplier":"—","sds_date":"—",
                "alert_level":"ERROR",
                "alert_justification":f"⚠️ Technical error: {error_reason[:200]}",
                "tox_comment":"","status":"Error","format":"—",
            })
            with st.expander("Technical details"):
                st.code(traceback.format_exc())

    # ── 6. Results summary ────────────────────────────────────────────────────
    if results:
        write_project_tab(wb, config, results)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        excel_bytes = buf.getvalue()
        fname = (f"SDS_Analysis_{config.get('client_name','').replace(' ','_')}"
                 f"_{datetime.date.today().strftime('%Y%m%d')}.xlsx")

        st.divider()
        st.subheader(f"4 — Preview — {len(results)} SDS analysed")
        st.caption("Summary view only. The full Excel report (with all extracted data and database results) is sent by email.")

        # Build HTML table
        rows_html = ""
        for i, r in enumerate(results):
            lvl = r["alert_level"]
            bg, fg, emoji = ALERT_COLORS.get(lvl, ("#888","#fff","⚠️"))
            badge = (f'<span style="background:{bg};color:{fg};padding:3px 10px;'
                     f'border-radius:4px;font-size:0.82em;font-weight:600;white-space:nowrap;">'
                     f'{emoji} {lvl}</span>')
            justif = r["alert_justification"]
            if r.get("tox_comment"):
                justif += f"\n\nToxicologist comment: {r['tox_comment']}"
            # Convert newlines to <br> for HTML
            justif_html = justif.replace("\n", "<br>")

            row_bg = "#FFF8F8" if lvl == "CRITICAL" else                      "#FFF5EE" if lvl == "MAJOR" else                      "#FFFEF0" if lvl == "MINOR" else                      "#F5FFF8" if lvl == "NONE" else "#F5F5F5"

            rows_html += f"""
            <tr style="background:{row_bg};vertical-align:top;">
              <td style="padding:10px 8px;text-align:center;color:#666;font-size:0.85em;">{i+1}</td>
              <td style="padding:10px 8px;font-size:0.82em;color:#444;word-break:break-all;">{r["sds_id"] or "—"}</td>
              <td style="padding:10px 8px;font-weight:600;color:#222;">{r["product"]}</td>
              <td style="padding:10px 8px;color:#444;">{r["supplier"] or "—"}</td>
              <td style="padding:10px 8px;color:#444;white-space:nowrap;">{r["sds_date"] or "—"}</td>
              <td style="padding:10px 8px;text-align:center;">{badge}</td>
              <td style="padding:10px 8px;font-size:0.85em;color:#333;line-height:1.5;">{justif_html}</td>
            </tr>"""

        st.markdown(f"""
        <table style="width:100%;border-collapse:collapse;border:1px solid #DCE8FF;border-radius:8px;overflow:hidden;">
          <thead>
            <tr style="background:#007AFF;color:#fff;">
              <th style="padding:10px 8px;width:40px;">#</th>
              <th style="padding:10px 8px;width:140px;">SDS ID</th>
              <th style="padding:10px 8px;width:180px;">Product</th>
              <th style="padding:10px 8px;width:140px;">Supplier</th>
              <th style="padding:10px 8px;width:100px;">SDS Date</th>
              <th style="padding:10px 8px;width:110px;">Alert Level</th>
              <th style="padding:10px 8px;">Alert Justification</th>
            </tr>
          </thead>
          <tbody>
            {rows_html}
          </tbody>
        </table>
        """, unsafe_allow_html=True)

        # Show errors separately
        for r in results:
            if r["alert_level"] == "ERROR":
                st.warning(f"⚠️ **{r['product']}** — {r['alert_justification']}")

        st.divider()

        # Emails
        notify_email = st.secrets.get("NOTIFY_EMAIL","elodie.saudrais@aconis.fr")
        summary_rows = "".join(
            f'<tr><td>{i+1}</td><td>{r["sds_id"] or "—"}</td><td>{r["product"]}</td>'
            f'<td>{r["supplier"] or "—"}</td><td>{r["sds_date"] or "—"}</td>'
            f'<td><b>{r["alert_level"]}</b></td></tr>'
            for i, r in enumerate(results)
        )

        # Email to client
        if client_email:
            client_html = f"""
            <div style="font-family:Segoe UI Light,Arial,sans-serif;color:#222;">
            <img src="data:image/png;base64,{LOGO_B64}" width="140"><br><br>
            <h2 style="color:#007AFF;">Rapport de screening biocompatibilité · Biocompatibility Screening Report</h2>
            <p>Bonjour / Dear {client_name or "Client"},</p>
            <p>Veuillez trouver ci-joint votre rapport d'analyse de biocompatibilité généré par ACONIS.<br>
            Please find attached your biocompatibility screening report generated by ACONIS.</p>
            <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%;">
            <tr style="background:#007AFF;color:#fff;">
              <th>#</th><th>SDS ID</th><th>Produit</th><th>Fournisseur</th><th>Date SDS</th><th>Niveau d'alerte</th>
            </tr>
            {summary_rows}
            </table>
            <br>
            <p style="color:#555;font-size:0.88em;">
            📊 Ce tableau est un résumé. Le fichier Excel joint contient le détail complet des données
            extraites et des résultats de consultation des 8 bases de données réglementaires.<br>
            This table is a summary. The attached Excel file contains the full details of extracted data
            and results from all 8 regulatory databases.
            </p>
            <br>
            <p style="color:#888;font-size:0.82em;">
            <i>Résultats préliminaires générés par IA — nécessitent une validation par un toxicologue qualifié.<br>
            Preliminary AI-assisted results — require validation by a qualified toxicologist.</i><br><br>
            <b>ACONIS SAS</b> — 6, rue Bouchard 69510 Messimy, France<br>
            RCS Lyon n° 944 426 113 — SIRET 944 426 113 00014<br>
            elodie.saudrais@aconis.fr — +33 6 11 38 53 65
            </p>
            </div>"""
            ok = send_email_with_attachment(
                client_email,
                f"Rapport biocompatibilité SDS — {client_name}",
                client_html, excel_bytes, fname)
            if ok:
                st.success(f"📧 Rapport Excel envoyé à **{client_email}**")

        # Notification to Élodie
        aconis_html = f"""
        <div style="font-family:Segoe UI Light,Arial,sans-serif;color:#222;">
        <h2 style="color:#007AFF;">Nouveau screening complété</h2>
        <p><b>Date :</b> {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
        <p><b>Client :</b> {client_name}</p>
        <p><b>Email client :</b> {client_email}</p>
        <p><b>Fichiers analysés :</b> {len(results)}</p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%;">
        <tr style="background:#007AFF;color:#fff;">
          <th>#</th><th>SDS ID</th><th>Produit</th><th>Fournisseur</th><th>Date SDS</th><th>Alerte</th>
        </tr>
        {summary_rows}
        </table>
        </div>"""
        ok2 = send_email_with_attachment(
            notify_email,
            f"[SDS Screener] {client_name} — {datetime.date.today()} — {len(results)} FDS",
            aconis_html, excel_bytes, fname)
        if ok2:
            st.success(f"📧 Résumé envoyé à **{notify_email}**")

        st.info("📊 The preview above is a **summary only**. "
                "The **Excel file sent by email** contains the full detail of all extracted data "
                "and results from all 8 regulatory databases.")

        st.download_button(
            label=f"📥 Télécharger {fname}",
            data=excel_bytes, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True,
        )

# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer-aconis">
    Aconis SAS — 6, rue Bouchard 69510 Messimy, France<br>
    RCS Lyon n° 944 426 113 — SIRET 944 426 113 00014<br>
    <a href="mailto:elodie.saudrais@aconis.fr" style="color:#007AFF;">elodie.saudrais@aconis.fr</a>
    &nbsp;·&nbsp; +33 6 11 38 53 65
</div>
""", unsafe_allow_html=True)
